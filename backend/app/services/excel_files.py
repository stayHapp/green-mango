"""会议嘉宾 Excel 模板、导入和签到导出服务。"""

from io import BytesIO
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import Session, aliased

from app.models.application import GuestApplication
from app.models.guest import CheckIn, Guest, GuestField, GuestValue
from app.models.meeting import Meeting
from app.models.user import User
from app.core.config import settings
from app.schemas.admin_resources import GuestImportResponse, ImportRowError
from app.schemas.guest import GuestCreate, GuestUpdate
from app.services.admin_guests import create_guest
from app.services.admin_resources import get_guest_registration_settings, update_guest
from app.services.check_in_sessions import get_default_check_in_session

FIXED_IMPORT_COLUMNS = {
    "姓名": "name",
    "手机号": "phone",
    "单位": "organization",
    "职务": "title",
    "身份": "tag",
    "座位号": "seat",
}
REQUIRED_IMPORT_COLUMNS = {"姓名"}
MAX_IMPORT_ROWS = 10_000
FIXED_EXPORT_COLUMNS = [
    ("name", "姓名", 18),
    ("phone", "手机号", 20),
    ("organization", "单位", 28),
    ("title", "职务", 20),
    ("tag", "身份", 16),
    ("seat", "座位号", 16),
]


def style_worksheet(worksheet, column_widths: list[int]) -> None:
    """为导入模板和导出表格设置统一的可读样式。

    入参：worksheet 为 openpyxl 工作表对象，必填；column_widths 为各列建议宽度，元素必须为正整数。
    返回值：None：直接修改传入工作表的表头、冻结窗格、筛选和列宽。
    异常：工作表对象不支持 openpyxl 接口或宽度参数无效时由 openpyxl 抛出异常。
    """
    header_fill = PatternFill(fill_type="solid", fgColor="2F6B4F")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def workbook_bytes(workbook: Workbook) -> bytes:
    """将内存工作簿序列化为标准 XLSX 字节。

    入参：workbook 为待输出的 openpyxl 工作簿，必填。
    返回值：bytes：可直接作为 HTTP 响应体或测试输入的 XLSX 内容。
    异常：工作簿包含不可序列化内容时由 openpyxl 抛出异常。
    """
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def display_export_phone(phone: str) -> str:
    """把默认手机号转换为导出的「未提供」文案。

    入参：phone 为嘉宾手机号，必填。
    返回值：str：等于默认手机号时返回“未提供”，否则返回原号码。
    异常：当前函数不主动抛出异常。
    """
    return "未提供" if phone == settings.default_guest_phone else phone


def build_guest_import_template(db: Session, meeting: Meeting) -> bytes:
    """生成包含固定字段和会议动态字段的嘉宾导入模板。

    入参：db 为数据库会话；meeting 为已完成管理员授权校验的会议，均必填。
    返回值：bytes：带表头、填写说明和示例数据的 XLSX 文件内容。
    异常：数据库查询或工作簿序列化失败时向上抛出对应异常。
    """
    fields = list(
        db.scalars(
            select(GuestField)
            .where(GuestField.meeting_id == meeting.id, GuestField.is_enabled.is_(True))
            .order_by(GuestField.sort_order, GuestField.id)
        )
    )
    dynamic_fields = fields
    extra_dynamic_fields = [field for field in dynamic_fields if field.label not in FIXED_IMPORT_COLUMNS]
    headers = [*FIXED_IMPORT_COLUMNS, *(field.label for field in extra_dynamic_fields)]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "嘉宾导入"
    worksheet.append(headers)
    example_values = ["张老师", "13800000000", "示例学校", "教研主任", "参会嘉宾", "A01"]
    example_values.extend("必填示例" if field.required else "" for field in extra_dynamic_fields)
    worksheet.append(example_values)
    style_worksheet(worksheet, [18, 20, 28, 20, 16, 16, *([20] * len(extra_dynamic_fields))])

    instruction = workbook.create_sheet("填写说明")
    instruction.append(["字段", "要求"])
    instruction.append(["姓名", "必填。"])
    instruction.append(["手机号", f"可选；列可省略或留空，省略/留空时自动使用默认手机号 {settings.default_guest_phone}。"])
    instruction.append(["单位、职务、身份、座位号", "可选。"])
    instruction.append(["表头", "顺序可任意；未识别的多余列会自动忽略，不影响导入。"])
    instruction.append(["数据行数", f"单次最多 {MAX_IMPORT_ROWS} 行，示例行可删除。"])
    for field in dynamic_fields:
        requirement = "必填" if field.required else "可选"
        instruction.append([field.label, f"会议自定义字段（{requirement}）"])
    style_worksheet(instruction, [24, 70])
    return workbook_bytes(workbook)


def normalize_cell(value: object) -> str | None:
    """将 Excel 单元格值规范化为业务文本。

    入参：value 为任意 openpyxl 单元格值，可为空、字符串、数字或日期。
    返回值：str | None：去除首尾空白的文本；空值或空白文本返回 None。
    异常：当前函数不主动抛出业务异常。
    """
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_import_rows(db: Session, meeting: Meeting, source: BinaryIO) -> GuestImportResponse:
    """解析并导入 Excel 第一张工作表中的嘉宾数据。

    入参：db 为数据库会话；meeting 为已授权会议；source 为可读取的 XLSX 二进制流，均必填。
    返回值：GuestImportResponse：成功导入数量和逐行错误信息；合法行会被保留，错误行不会写入。
    异常：文件不是有效 XLSX、缺少必填表头、表头重复或超过最大行数时抛出 ValueError。
    使用示例：路由将 UploadFile.file 传入本函数，并把返回摘要直接序列化为 JSON。
    """
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as error:
        raise ValueError("无法读取 Excel 文件，请确认文件为有效的 .xlsx 格式。") from error
    try:
        worksheet = workbook.worksheets[0]
        raw_headers = [normalize_cell(cell.value) for cell in worksheet[1]]
        headers = [header or "" for header in raw_headers]
        nonempty_headers = [header for header in headers if header]
        if len(nonempty_headers) != len(set(nonempty_headers)):
            raise ValueError("Excel 表头不能重复。")
        missing_headers = REQUIRED_IMPORT_COLUMNS - set(headers)
        if missing_headers:
            raise ValueError(f"Excel 缺少必填表头：{', '.join(sorted(missing_headers))}。")

        fields = list(
            db.scalars(
                select(GuestField).where(
                    GuestField.meeting_id == meeting.id,
                    GuestField.is_enabled.is_(True),
                )
            )
        )
        dynamic_by_label = {field.label: field for field in fields}
        if max(worksheet.max_row - 1, 0) > MAX_IMPORT_ROWS:
            raise ValueError(f"单次最多导入 {MAX_IMPORT_ROWS} 行嘉宾数据。")

        imported_count = 0
        updated_count = 0
        default_phone_count = 0
        errors: list[ImportRowError] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            values_by_header = {
                header: normalize_cell(row[index]) if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
            # 完全空白的行不计入错误，也不创建嘉宾。
            if not any(values_by_header.values()):
                continue
            # 姓名必填；手机号允许留空，留空时自动填充默认手机号。
            missing_values = [header for header in ("姓名",) if not values_by_header.get(header)]
            if missing_values:
                errors.append(ImportRowError(row_number=row_number, message=f"缺少必填值：{', '.join(sorted(missing_values))}。"))
                continue
            raw_fixed_values = {
                field_name: values_by_header.get(header)
                for header, field_name in FIXED_IMPORT_COLUMNS.items()
            }
            used_default_phone = False
            if not raw_fixed_values.get("phone"):
                raw_fixed_values["phone"] = settings.default_guest_phone
                used_default_phone = True
            fixed_values = raw_fixed_values
            dynamic_values = {
                field.key: values_by_header.get(label)
                for label, field in dynamic_by_label.items()
                if label in values_by_header
            }
            try:
                existing = db.scalar(
                    select(Guest).where(
                        Guest.meeting_id == meeting.id,
                        Guest.name == fixed_values["name"],
                        Guest.phone == fixed_values["phone"],
                        Guest.is_active.is_(True),
                    )
                )
                if existing is not None:
                    # 同会议同名同手机号的启用嘉宾视为已存在，覆盖更新其他字段并保留签到记录。
                    update_guest(db, meeting, existing, GuestUpdate(**fixed_values, values=dynamic_values))
                    updated_count += 1
                else:
                    create_guest(
                        db,
                        meeting,
                        GuestCreate(**fixed_values, values=dynamic_values),
                        source="admin_import",
                    )
                    imported_count += 1
                if used_default_phone:
                    default_phone_count += 1
            except (ValidationError, ValueError, RuntimeError) as error:
                db.rollback()
                errors.append(ImportRowError(row_number=row_number, message=str(error)))
        return GuestImportResponse(
            imported_count=imported_count,
            updated_count=updated_count,
            default_phone_count=default_phone_count,
            errors=errors,
        )
    finally:
        workbook.close()


def build_check_in_export(db: Session, meeting: Meeting) -> bytes:
    """导出会议完整嘉宾名单及签到结果。

    入参：db 为数据库会话；meeting 为已完成管理员授权校验的会议，均必填。
    返回值：bytes：包含每位嘉宾嘉宾类型、陪同嘉宾、签到状态、时间、方式和工作人员的 XLSX 内容。
    异常：数据库查询或工作簿序列化失败时向上抛出对应异常。
    """
    default_session = get_default_check_in_session(db, meeting)
    _, _, enabled_fixed_fields = get_guest_registration_settings(meeting)
    fixed_columns = [column for column in FIXED_EXPORT_COLUMNS if column[0] in enabled_fixed_fields]
    primary_alias = aliased(Guest)
    companion_primary_names = dict(
        db.execute(
            select(Guest.id, primary_alias.name)
            .join(primary_alias, primary_alias.id == Guest.companion_of_id)
            .where(Guest.meeting_id == meeting.id)
        ).all()
    )
    statement = (
        select(Guest, CheckIn, func.coalesce(User.display_name, User.username))
        .outerjoin(CheckIn, (CheckIn.session_id == default_session.id) & (CheckIn.guest_id == Guest.id))
        .outerjoin(User, User.id == CheckIn.staff_id)
        .where(Guest.meeting_id == meeting.id, Guest.is_active.is_(True))
        .order_by(Guest.created_at, Guest.id)
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "签到明细"
    worksheet.append([
        "嘉宾ID",
        *(label for _, label, _ in fixed_columns),
        "嘉宾类型",
        "陪同嘉宾",
        "签到状态",
        "签到时间",
        "签到方式",
        "执行工作人员",
    ])
    for guest, check_in, staff_name in db.execute(statement).tuples():
        worksheet.append([
            guest.id,
            *(display_export_phone(getattr(guest, key)) if key == "phone" else getattr(guest, key) for key, _, _ in fixed_columns),
            "同行人员" if guest.companion_of_id is not None else "本人",
            companion_primary_names.get(guest.id) or "",
            "已签到" if check_in else "未签到",
            check_in.checked_in_at.isoformat() if check_in else None,
            check_in.method if check_in else None,
            staff_name,
        ])
    style_worksheet(worksheet, [12, *(width for _, _, width in fixed_columns), 14, 18, 14, 28, 16, 22])
    return workbook_bytes(workbook)


def build_guest_status_export(db: Session, meeting: Meeting) -> bytes:
    """导出与嘉宾管理列表口径一致的嘉宾信息和状态表。

    入参：db 为数据库会话；meeting 为已完成管理员授权校验的会议，均必填。
    返回值：bytes：包含正式嘉宾、待审核和已拒绝报名申请，以及来源、管理状态、签到状态的 XLSX 内容。
    异常：数据库查询或工作簿序列化失败时向上抛出对应异常。
    使用示例：`content = build_guest_status_export(db, meeting)`。
    """
    fields = list(
        db.scalars(
            select(GuestField)
            .where(GuestField.meeting_id == meeting.id, GuestField.is_enabled.is_(True))
            .order_by(GuestField.sort_order, GuestField.id)
        )
    )
    extra_fields = [field for field in fields if field.label not in FIXED_IMPORT_COLUMNS]
    _, _, enabled_fixed_fields = get_guest_registration_settings(meeting)
    fixed_columns = [column for column in FIXED_EXPORT_COLUMNS if column[0] in enabled_fixed_fields]
    guest_values = {
        (value.guest_id, value.field_key): value.value_text
        for value in db.scalars(
            select(GuestValue)
            .join(Guest, Guest.id == GuestValue.guest_id)
            .where(Guest.meeting_id == meeting.id, Guest.is_active.is_(True))
        )
    }
    primary_alias = aliased(Guest)
    companion_primary_names = dict(
        db.execute(
            select(Guest.id, primary_alias.name)
            .join(primary_alias, primary_alias.id == Guest.companion_of_id)
            .where(Guest.meeting_id == meeting.id)
        ).all()
    )
    default_session = get_default_check_in_session(db, meeting)
    guest_statement = (
        select(Guest, CheckIn)
        .outerjoin(CheckIn, (CheckIn.session_id == default_session.id) & (CheckIn.guest_id == Guest.id))
        .where(Guest.meeting_id == meeting.id, Guest.is_active.is_(True))
        .order_by(Guest.created_at, Guest.id)
    )
    application_statement = (
        select(GuestApplication)
        .where(
            GuestApplication.meeting_id == meeting.id,
            GuestApplication.status != "approved",
        )
        .order_by(GuestApplication.created_at, GuestApplication.id)
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "嘉宾状态"
    worksheet.append([
        "记录ID",
        *(label for _, label, _ in fixed_columns),
        *(field.label for field in extra_fields),
        "嘉宾类型",
        "陪同嘉宾",
        "来源",
        "管理状态",
        "签到状态",
    ])

    source_labels = {
        "self_registration": "自主报名",
        "admin_import": "后台导入",
        "admin_entry": "后台录入",
        "companion_registration": "同行登记",
    }
    # 未转化的报名申请先输出，与嘉宾管理列表的排列口径保持一致。
    for application in db.scalars(application_statement):
        worksheet.append([
            f"申请-{application.id}",
            *(display_export_phone(getattr(application, key)) if key == "phone" else getattr(application, key) for key, _, _ in fixed_columns),
            *(application.values_json.get(field.key) for field in extra_fields),
            "—",
            "—",
            "自主报名",
            "待审核" if application.status == "pending" else "已拒绝",
            "—",
        ])

    for guest, check_in in db.execute(guest_statement).tuples():
        worksheet.append([
            f"嘉宾-{guest.id}",
            *(display_export_phone(getattr(guest, key)) if key == "phone" else getattr(guest, key) for key, _, _ in fixed_columns),
            *(guest_values.get((guest.id, field.key)) for field in extra_fields),
            "同行人员" if guest.companion_of_id is not None else "本人",
            companion_primary_names.get(guest.id) or "",
            source_labels.get(guest.source, "后台录入"),
            "已通过" if guest.source == "self_registration" else "已录入",
            "已签到" if check_in else "未签到",
        ])

    column_widths = [16, *(width for _, _, width in fixed_columns), *([20] * len(extra_fields)), 14, 18, 16, 14, 14]
    style_worksheet(worksheet, column_widths)
    return workbook_bytes(workbook)
