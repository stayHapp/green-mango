"""管理员会议管理 API 测试。"""

from collections.abc import Generator
from datetime import datetime, timedelta, timezone
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

import app.models  # noqa: F401  # 导入模型以注册完整 metadata（元数据）。
from app.api.dependencies import get_db
from app.core.security import hash_password
from app.db import Base
from app.main import create_app
from app.models.access import MeetingAdmin, StaffMeeting
from app.models.application import GuestApplication
from app.models.guest import CheckIn, Guest, GuestField, GuestValue
from app.models.meeting import Meeting, MeetingSetting
from app.models.user import User
from app.services.sessions import create_guest_session, create_user_session


@pytest.fixture
def client_and_session(tmp_path) -> Generator[tuple[TestClient, Session], None, None]:
    """创建隔离数据库、测试客户端和请求依赖覆盖。

    入参：tmp_path 为 pytest 临时目录，用于创建隔离 SQLite 数据库。
    返回值：Generator[tuple[TestClient, Session], None, None]：测试客户端与可用于准备数据的数据库会话。
    异常：数据库建表或测试客户端初始化失败时，由 SQLAlchemy 或 FastAPI 抛出异常。
    """
    engine = create_engine(f"sqlite:///{tmp_path / 'admin-meetings.db'}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    app = create_app()

    def override_get_db() -> Generator[Session, None, None]:
        """为每个测试请求提供独立数据库会话。

        入参：无。
        返回值：Generator[Session, None, None]：测试专用数据库会话。
        异常：会话创建失败时由 SQLAlchemy 抛出异常。
        """
        db = session_factory()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    db = session_factory()
    try:
        with TestClient(app) as client:
            yield client, db
    finally:
        db.close()
        app.dependency_overrides.clear()
        engine.dispose()


def create_user(
    db: Session,
    username: str,
    role: str = "admin",
    is_active: bool = True,
    password: str | None = None,
) -> User:
    """创建用于管理员 API 测试的用户。

    入参：db 为数据库会话；username 为唯一账号；role 为角色；is_active 表示账号是否启用。
    返回值：User：已持久化并具有主键的用户对象。
    异常：账号重复或数据库写入失败时由 SQLAlchemy 抛出异常。
    """
    user = User(
        username=username,
        password_hash=hash_password(password) if password else "test-hash",
        display_name=username,
        role=role,
        is_active=is_active,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def auth_headers(db: Session, subject: User | Guest) -> dict[str, str]:
    """为测试用户或嘉宾创建真实 Bearer 会话请求头。

    入参：db 为数据库会话；subject 为管理员、工作人员或嘉宾对象，均必填。
    返回值：dict[str, str]：包含 Authorization Bearer token 的请求头。
    异常：会话写入失败时由 SQLAlchemy 抛出异常。
    """
    if isinstance(subject, Guest):
        token, _ = create_guest_session(db, subject)
    else:
        token, _ = create_user_session(db, subject)
    return {"Authorization": f"Bearer {token}"}


def test_admin_can_create_list_get_and_update_meeting(client_and_session: tuple[TestClient, Session]) -> None:
    """验证管理员可完成会议创建、列表、详情和修改闭环。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示会议管理核心路径可用。
    异常：当前函数不主动抛出业务异常；断言失败表示接口行为不符合预期。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-one")
    headers = auth_headers(db, admin)

    create_response = client.post(
        "/api/admin/meetings",
        headers=headers,
        json={
            "title": "2026 教育创新论坛",
            "description": "会议说明",
            "location": "杭州",
            "start_time": "2026-08-01T09:00:00+08:00",
            "end_time": "2026-08-01T17:00:00+08:00",
            "status": "draft",
        },
    )
    assert create_response.status_code == 201
    meeting_id = create_response.json()["id"]

    list_response = client.get("/api/admin/meetings", headers=headers)
    assert list_response.status_code == 200
    assert [item["id"] for item in list_response.json()] == [meeting_id]

    detail_response = client.get(f"/api/admin/meetings/{meeting_id}", headers=headers)
    assert detail_response.status_code == 200
    assert detail_response.json()["title"] == "2026 教育创新论坛"

    update_response = client.patch(
        f"/api/admin/meetings/{meeting_id}", headers=headers, json={"status": "published"}
    )
    assert update_response.status_code == 200
    assert update_response.json()["status"] == "published"
    assert db.query(MeetingAdmin).filter_by(meeting_id=meeting_id, user_id=admin.id).count() == 1


def test_admin_cannot_access_unassigned_meeting(client_and_session: tuple[TestClient, Session]) -> None:
    """验证管理员无法读取未被授权的会议。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示会议授权过滤生效。
    异常：当前函数不主动抛出业务异常；断言失败表示存在越权风险。
    """
    client, db = client_and_session
    owner = create_user(db, "admin-owner")
    visitor = create_user(db, "admin-visitor")
    meeting = Meeting(title="受限会议", created_by_id=owner.id, status="draft")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=owner.id))
    db.commit()

    response = client.get(f"/api/admin/meetings/{meeting.id}", headers=auth_headers(db, visitor))
    assert response.status_code == 404


def test_non_admin_or_disabled_user_cannot_use_admin_api(client_and_session: tuple[TestClient, Session]) -> None:
    """验证工作人员和已停用管理员不能访问管理员接口。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示过渡期管理员依赖完成角色与启用状态校验。
    异常：当前函数不主动抛出业务异常；断言失败表示权限校验存在缺口。
    """
    client, db = client_and_session
    staff = create_user(db, "staff-one", role="staff")
    disabled_admin = create_user(db, "admin-disabled", is_active=False)

    staff_response = client.get("/api/admin/meetings", headers=auth_headers(db, staff))
    disabled_response = client.get("/api/admin/meetings", headers=auth_headers(db, disabled_admin))

    assert staff_response.status_code == 403
    assert disabled_response.status_code == 403


def test_admin_login_and_logout_use_revocable_bearer_session(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证管理员密码登录、Bearer 鉴权和退出撤销会话。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示安全会话完整闭环可用。
    异常：当前函数不主动抛出业务异常；断言失败表示认证或撤销逻辑异常。
    """
    client, db = client_and_session
    create_user(db, "secure-admin", password="safe-password-123")

    login_response = client.post(
        "/api/admin/sessions", json={"username": "secure-admin", "password": "safe-password-123"}
    )
    assert login_response.status_code == 200
    headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}
    assert client.get("/api/admin/meetings", headers=headers).status_code == 200
    assert client.post("/api/sessions/logout", headers=headers).status_code == 200
    assert client.get("/api/admin/meetings", headers=headers).status_code == 401


def test_meeting_time_range_is_validated(client_and_session: tuple[TestClient, Session]) -> None:
    """验证创建和修改会议时拒绝无效的起止时间。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示会议时间范围校验生效。
    异常：当前函数不主动抛出业务异常；断言失败表示无效时间可能进入数据库。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-time")
    headers = auth_headers(db, admin)

    invalid_create = client.post(
        "/api/admin/meetings",
        headers=headers,
        json={"title": "无效会议", "start_time": "2026-08-02T10:00:00+08:00", "end_time": "2026-08-02T09:00:00+08:00"},
    )
    assert invalid_create.status_code == 422

    meeting = Meeting(title="有效会议", created_by_id=admin.id, status="draft")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.commit()
    invalid_update = client.patch(
        f"/api/admin/meetings/{meeting.id}",
        headers=headers,
        json={"start_time": "2026-08-02T10:00:00+08:00", "end_time": "2026-08-02T09:00:00+08:00"},
    )
    assert invalid_update.status_code == 422


def test_admin_can_replace_guest_fields_and_create_guest(client_and_session: tuple[TestClient, Session]) -> None:
    """验证管理员可配置嘉宾字段并录入带二维码凭证的嘉宾。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示嘉宾字段与嘉宾录入接口主路径可用。
    异常：当前函数不主动抛出业务异常；断言失败表示接口行为不符合预期。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-guests")
    meeting = Meeting(title="嘉宾管理会议", created_by_id=admin.id, status="draft")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.commit()
    headers = auth_headers(db, admin)

    fields_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-fields",
        headers=headers,
        json={"fields": [{"label": "单位", "key": "organization", "field_type": "text", "sort_order": 1}]},
    )
    assert fields_response.status_code == 200
    assert fields_response.json()[0]["key"] == "organization"

    create_response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests",
        headers=headers,
        json={"name": "李文博", "phone": "13900000001", "organization": "知会教育", "tag": "参会嘉宾"},
    )
    assert create_response.status_code == 201
    assert create_response.json()["name"] == "李文博"
    assert create_response.json()["qr_token"]

    list_response = client.get(f"/api/admin/meetings/{meeting.id}/guests", headers=headers)
    assert list_response.status_code == 200
    assert len(list_response.json()) == 1


def test_guest_field_save_updates_in_place_and_protects_destructive_changes(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证字段配置增量更新已有字段，并保护含值字段免受删除或类型变更。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示安全修改保留字段 ID 和嘉宾值，破坏性修改会被拒绝。
    异常：当前函数不主动抛出业务异常；断言失败表示数据一致性保护失效。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-fields")
    meeting = Meeting(title="字段保护会议", created_by_id=admin.id, status="draft")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.commit()
    headers = auth_headers(db, admin)

    duplicate_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-fields",
        headers=headers,
        json={
            "fields": [
                {"label": "单位", "key": "organization", "field_type": "text"},
                {"label": "单位副本", "key": "organization", "field_type": "text"},
            ]
        },
    )
    assert duplicate_response.status_code == 422

    field = GuestField(meeting_id=meeting.id, label="单位", key="organization", field_type="text")
    guest = Guest(meeting_id=meeting.id, name="王敏", phone="13900000002", qr_token="test-guest-token")
    db.add_all([field, guest])
    db.flush()
    original_field_id = field.id
    db.add(GuestValue(guest_id=guest.id, field_id=field.id, field_key=field.key, value_text="知会教育"))
    db.commit()

    update_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-fields",
        headers=headers,
        json={
            "fields": [
                {
                    "label": "所在单位",
                    "key": "organization",
                    "field_type": "text",
                    "required": True,
                    "visible_to_guest": False,
                    "is_enabled": True,
                    "sort_order": 1,
                },
                {
                    "label": "饮食偏好",
                    "key": "diet_preference",
                    "field_type": "text",
                    "sort_order": 2,
                },
            ]
        },
    )
    assert update_response.status_code == 200
    assert update_response.json()[0]["id"] == original_field_id
    assert update_response.json()[0]["label"] == "所在单位"
    assert update_response.json()[0]["required"] is True
    saved_value = db.scalar(select(GuestValue).where(GuestValue.guest_id == guest.id))
    assert saved_value is not None
    assert saved_value.field_id == original_field_id
    assert saved_value.value_text == "知会教育"

    empty_field = db.scalar(
        select(GuestField).where(
            GuestField.meeting_id == meeting.id,
            GuestField.key == "diet_preference",
        )
    )
    assert empty_field is not None
    empty_field_id = empty_field.id
    db.add(GuestValue(guest_id=guest.id, field_id=empty_field.id, field_key=empty_field.key, value_text=""))
    db.commit()

    remove_empty_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-fields",
        headers=headers,
        json={
            "fields": [
                {
                    "label": "所在单位",
                    "key": "organization",
                    "field_type": "text",
                }
            ]
        },
    )
    assert remove_empty_response.status_code == 200
    assert [item["key"] for item in remove_empty_response.json()] == ["organization"]
    assert db.scalar(select(GuestValue).where(GuestValue.field_id == empty_field_id)) is None

    delete_valued_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-fields", headers=headers, json={"fields": []}
    )
    assert delete_valued_response.status_code == 422
    assert delete_valued_response.json()["detail"] == "字段“所在单位”已有嘉宾数据，不能删除。"

    change_type_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-fields",
        headers=headers,
        json={
            "fields": [
                {
                    "label": "所在单位",
                    "key": "organization",
                    "field_type": "select",
                }
            ]
        },
    )
    assert change_type_response.status_code == 422
    assert change_type_response.json()["detail"] == "字段“所在单位”已有嘉宾数据，不能修改字段类型。"


def test_admin_configures_guest_display_fields_and_guest_profile_receives_them(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证管理员可选择固定与动态嘉宾呈现字段，并由嘉宾资料接口读取。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示会议级呈现配置保留顺序、校验 key 并传递动态字段标签。
    异常：当前函数不主动抛出业务异常；断言失败表示管理员配置或嘉宾呈现契约不一致。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-display-fields")
    meeting = Meeting(title="字段呈现会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add_all(
        [
            MeetingAdmin(meeting_id=meeting.id, user_id=admin.id),
            MeetingSetting(meeting_id=meeting.id, settings_json={}),
            GuestField(
                meeting_id=meeting.id,
                label="饮食偏好",
                key="diet_preference",
                field_type="text",
                visible_to_guest=True,
            ),
        ]
    )
    guest = Guest(
        meeting_id=meeting.id,
        name="周老师",
        phone="13900000042",
        organization="省教育科学研究院",
        title="研究员",
        tag="主讲嘉宾",
        seat="A42",
        qr_token="display-fields-token",
    )
    db.add(guest)
    db.commit()

    admin_headers = auth_headers(db, admin)
    default_response = client.get(
        f"/api/admin/meetings/{meeting.id}/guest-display-fields",
        headers=admin_headers,
    )
    assert default_response.status_code == 200
    assert default_response.json()["fields"] == [
        "name",
        "phone",
        "organization",
        "title",
        "tag",
        "seat",
        "diet_preference",
    ]

    save_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-display-fields",
        headers=admin_headers,
        json={"fields": ["name", "organization", "seat", "diet_preference"]},
    )
    assert save_response.status_code == 200
    assert save_response.json()["fields"] == ["name", "organization", "seat", "diet_preference"]

    invalid_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-display-fields",
        headers=admin_headers,
        json={"fields": ["unknown_field"]},
    )
    assert invalid_response.status_code == 422

    profile_response = client.get(
        f"/api/guest/meetings/{meeting.id}/profile",
        headers=auth_headers(db, guest),
    )
    assert profile_response.status_code == 200
    assert profile_response.json()["visible_fields"] == ["name", "organization", "seat", "diet_preference"]
    assert profile_response.json()["field_labels"] == {"diet_preference": "饮食偏好"}


def test_guest_endpoints_reject_unassigned_admin(client_and_session: tuple[TestClient, Session]) -> None:
    """验证未获会议授权的管理员不能访问嘉宾资源。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示嘉宾资源沿用会议授权保护。
    异常：当前函数不主动抛出业务异常；断言失败表示存在越权风险。
    """
    client, db = client_and_session
    owner = create_user(db, "admin-guest-owner")
    visitor = create_user(db, "admin-guest-visitor")
    meeting = Meeting(title="受限嘉宾会议", created_by_id=owner.id, status="draft")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=owner.id))
    db.commit()

    response = client.get(f"/api/admin/meetings/{meeting.id}/guests", headers=auth_headers(db, visitor))
    assert response.status_code == 404


def test_admin_can_create_staff_and_staff_can_list_assigned_meetings(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证管理员创建工作人员后，工作人员可查询负责会议。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示工作人员授权链路可用。
    异常：当前函数不主动抛出业务异常；断言失败表示创建或授权逻辑不符合预期。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-staff-create")
    meeting = Meeting(title="工作人员会议", created_by_id=admin.id, status="draft")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.commit()
    admin_headers = auth_headers(db, admin)

    create_response = client.post(
        f"/api/admin/meetings/{meeting.id}/staff",
        headers=admin_headers,
        json={"username": "staff01", "display_name": "现场一组", "phone": "13700000001", "initial_password": "staff-pass-123"},
    )
    assert create_response.status_code == 201
    staff_id = create_response.json()["id"]

    repeated_response = client.post(
        f"/api/admin/meetings/{meeting.id}/staff",
        headers=admin_headers,
        json={"username": "staff01", "display_name": "现场一组", "initial_password": "staff-pass-123"},
    )
    assert repeated_response.status_code == 201
    assert repeated_response.json()["id"] == staff_id

    login_response = client.post(
        "/api/staff/sessions", json={"username": "staff01", "password": "staff-pass-123"}
    )
    assert login_response.status_code == 200

    list_response = client.get(f"/api/admin/meetings/{meeting.id}/staff", headers=admin_headers)
    assert list_response.status_code == 200
    assert [item["id"] for item in list_response.json()] == [staff_id]

    db.expire_all()
    staff_user = db.get(User, staff_id)
    assert staff_user is not None
    staff_meetings_response = client.get("/api/staff/meetings", headers=auth_headers(db, staff_user))
    assert staff_meetings_response.status_code == 200
    assert [item["id"] for item in staff_meetings_response.json()] == [meeting.id]


def test_staff_meetings_reject_non_staff_user(client_and_session: tuple[TestClient, Session]) -> None:
    """验证管理员账号不能伪装为工作人员读取负责会议。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示工作人员角色校验生效。
    异常：当前函数不主动抛出业务异常；断言失败表示角色权限存在缺口。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-not-staff")
    response = client.get("/api/staff/meetings", headers=auth_headers(db, admin))
    assert response.status_code == 403


def test_staff_can_scan_and_manually_check_in_with_core_rules(client_and_session: tuple[TestClient, Session]) -> None:
    """验证工作人员扫码、人工签到与重复签到保护。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示工作人员签到核心链路可用。
    异常：当前函数不主动抛出业务异常；断言失败表示签到规则不符合预期。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-check-in")
    staff = create_user(db, "staff-check-in", role="staff")
    meeting = Meeting(
        title="签到会议",
        created_by_id=admin.id,
        status="published",
        end_time=datetime.now(timezone.utc) + timedelta(days=1),
    )
    db.add(meeting)
    db.flush()
    db.add(StaffMeeting(meeting_id=meeting.id, user_id=staff.id))
    first_guest = Guest(meeting_id=meeting.id, name="扫码嘉宾", phone="13900000003", qr_token="scan-token")
    second_guest = Guest(meeting_id=meeting.id, name="人工嘉宾", phone="13900000004", qr_token="manual-token")
    db.add_all([first_guest, second_guest])
    db.commit()
    headers = auth_headers(db, staff)

    scan_response = client.post(
        f"/api/staff/meetings/{meeting.id}/check-ins/scan", headers=headers, json={"qr_token": "scan-token"}
    )
    assert scan_response.status_code == 201
    assert scan_response.json()["method"] == "scan"

    duplicate_response = client.post(
        f"/api/staff/meetings/{meeting.id}/check-ins/scan", headers=headers, json={"qr_token": "scan-token"}
    )
    assert duplicate_response.status_code == 409

    manual_response = client.post(
        f"/api/staff/meetings/{meeting.id}/check-ins/manual", headers=headers, json={"guest_id": second_guest.id}
    )
    assert manual_response.status_code == 201
    assert manual_response.json()["method"] == "manual"

    records_response = client.get(f"/api/staff/meetings/{meeting.id}/check-ins", headers=headers)
    assert records_response.status_code == 200
    assert len(records_response.json()) == 2
    assert db.query(CheckIn).filter_by(meeting_id=meeting.id).count() == 2


def test_check_in_rejects_expired_or_other_meeting_guest(client_and_session: tuple[TestClient, Session]) -> None:
    """验证签到拒绝过期会议和其他会议嘉宾二维码。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示二维码有效范围校验生效。
    异常：当前函数不主动抛出业务异常；断言失败表示可能发生跨会议或过期签到。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-check-in-boundary")
    staff = create_user(db, "staff-check-in-boundary", role="staff")
    expired_meeting = Meeting(title="已结束会议", created_by_id=admin.id, status="ended", end_time=datetime.now(timezone.utc) - timedelta(days=1))
    active_meeting = Meeting(title="进行中会议", created_by_id=admin.id, status="published", end_time=datetime.now(timezone.utc) + timedelta(days=1))
    other_meeting = Meeting(title="其他会议", created_by_id=admin.id, status="published", end_time=datetime.now(timezone.utc) + timedelta(days=1))
    db.add_all([expired_meeting, active_meeting, other_meeting])
    db.flush()
    db.add_all([StaffMeeting(meeting_id=expired_meeting.id, user_id=staff.id), StaffMeeting(meeting_id=active_meeting.id, user_id=staff.id)])
    db.add_all([
        Guest(meeting_id=expired_meeting.id, name="过期嘉宾", phone="13900000005", qr_token="expired-token"),
        Guest(meeting_id=other_meeting.id, name="跨会嘉宾", phone="13900000006", qr_token="other-token"),
    ])
    db.commit()
    headers = auth_headers(db, staff)

    expired_response = client.post(f"/api/staff/meetings/{expired_meeting.id}/check-ins/scan", headers=headers, json={"qr_token": "expired-token"})
    other_response = client.post(f"/api/staff/meetings/{active_meeting.id}/check-ins/scan", headers=headers, json={"qr_token": "other-token"})
    assert expired_response.status_code == 422
    assert other_response.status_code == 422


def test_guest_can_login_and_read_own_meeting_and_qr(client_and_session: tuple[TestClient, Session]) -> None:
    """验证嘉宾可登录并读取自己的会议和二维码 token。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示嘉宾端 API 核心路径可用。
    异常：当前函数不主动抛出业务异常；断言失败表示登录或访问控制异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-guest-session")
    meeting = Meeting(title="嘉宾会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    guest = Guest(meeting_id=meeting.id, name="李文博", phone="13900000007", qr_token="guest-session-token")
    db.add(guest)
    db.commit()

    login_response = client.post("/api/guest/sessions", json={"meeting_id": meeting.id, "name": "李文博", "phone": "13900000007"})
    assert login_response.status_code == 200
    assert login_response.json()["guest_id"] == guest.id
    headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}

    meetings_response = client.get("/api/guest/meetings", headers=headers)
    detail_response = client.get(f"/api/guest/meetings/{meeting.id}", headers=headers)
    profile_response = client.get(f"/api/guest/meetings/{meeting.id}/profile", headers=headers)
    qr_response = client.get(f"/api/guest/meetings/{meeting.id}/check-in-qr", headers=headers)
    assert meetings_response.status_code == 200
    assert detail_response.status_code == 200
    assert profile_response.status_code == 200
    assert profile_response.json()["name"] == "李文博"
    assert qr_response.json() == {
        "qr_token": "guest-session-token",
        "expires_at": None,
        "is_checked_in": False,
        "checked_in_at": None,
    }

    checked_in_at = datetime.now(timezone.utc)
    db.add(
        CheckIn(
            meeting_id=meeting.id,
            guest_id=guest.id,
            staff_id=None,
            method="manual",
            checked_in_at=checked_in_at,
        )
    )
    db.commit()

    checked_in_qr_response = client.get(
        f"/api/guest/meetings/{meeting.id}/check-in-qr", headers=headers
    )
    assert checked_in_qr_response.status_code == 200
    assert checked_in_qr_response.json()["is_checked_in"] is True
    assert checked_in_qr_response.json()["checked_in_at"] == checked_in_at.isoformat().replace("+00:00", "Z")


def test_guest_login_and_cross_meeting_access_are_rejected(client_and_session: tuple[TestClient, Session]) -> None:
    """验证嘉宾登录失败和跨会议访问被拒绝。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示嘉宾身份与会议归属校验生效。
    异常：当前函数不主动抛出业务异常；断言失败表示存在越权风险。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-guest-boundary")
    first_meeting = Meeting(title="我的会议", created_by_id=admin.id, status="published")
    other_meeting = Meeting(title="其他会议", created_by_id=admin.id, status="published")
    db.add_all([first_meeting, other_meeting])
    db.flush()
    guest = Guest(meeting_id=first_meeting.id, name="王敏", phone="13900000008", qr_token="guest-boundary-token")
    db.add(guest)
    db.commit()

    failed_login = client.post("/api/guest/sessions", json={"meeting_id": first_meeting.id, "name": "王敏", "phone": "错误"})
    cross_access = client.get(f"/api/guest/meetings/{other_meeting.id}", headers=auth_headers(db, guest))
    assert failed_login.status_code == 401
    assert cross_access.status_code == 404


def test_staff_can_search_guests_with_check_in_status(client_and_session: tuple[TestClient, Session]) -> None:
    """验证工作人员可按现场字段搜索嘉宾并核验签到状态。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示搜索字段与签到状态均正确返回。
    异常：当前函数不主动抛出业务异常；断言失败表示现场核验信息不完整。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-search")
    staff = create_user(db, "staff-search", role="staff")
    meeting = Meeting(title="搜索会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(StaffMeeting(meeting_id=meeting.id, user_id=staff.id))
    guest = Guest(meeting_id=meeting.id, name="陈老师", phone="13900000009", organization="知会学校", seat="A12", qr_token="search-token")
    db.add(guest)
    db.flush()
    db.add(CheckIn(meeting_id=meeting.id, guest_id=guest.id, staff_id=staff.id, method="scan"))
    db.commit()
    headers = auth_headers(db, staff)

    response = client.get(f"/api/staff/meetings/{meeting.id}/guests?query=A12", headers=headers)
    assert response.status_code == 200
    assert response.json()[0]["name"] == "陈老师"
    assert response.json()[0]["checked_in"] is True


def test_admin_can_view_check_in_summary(client_and_session: tuple[TestClient, Session]) -> None:
    """验证管理员可查看会议签到统计与工作人员明细。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示管理员签到统计接口可用。
    异常：当前函数不主动抛出业务异常；断言失败表示统计或明细数据不正确。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-summary")
    staff = create_user(db, "staff-summary", role="staff")
    meeting = Meeting(title="统计会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    first_guest = Guest(meeting_id=meeting.id, name="已签到嘉宾", phone="13900000010", qr_token="summary-one")
    second_guest = Guest(meeting_id=meeting.id, name="未签到嘉宾", phone="13900000011", qr_token="summary-two")
    db.add_all([first_guest, second_guest])
    db.flush()
    db.add(CheckIn(meeting_id=meeting.id, guest_id=first_guest.id, staff_id=staff.id, method="scan"))
    db.commit()

    response = client.get(f"/api/admin/meetings/{meeting.id}/check-ins", headers=auth_headers(db, admin))
    assert response.status_code == 200
    assert response.json()["total_guests"] == 2
    assert response.json()["checked_in_count"] == 1
    assert response.json()["unchecked_count"] == 1
    assert response.json()["records"][0]["staff_name"] == "staff-summary"
    assert response.json()["records"][0]["checked_in_at"].endswith("Z")


def test_admin_excel_template_import_and_check_in_export(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证 Excel 模板、逐行导入和完整签到导出闭环。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示生成和解析的文件均为真实可读 XLSX。
    异常：当前函数不主动抛出业务异常；断言失败表示 Excel API 行为不符合预期。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-excel")
    staff = create_user(db, "staff-excel", role="staff")
    meeting = Meeting(title="Excel 验证会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.add(GuestField(meeting_id=meeting.id, label="学科", key="subject", field_type="text", required=True))
    db.commit()
    headers = auth_headers(db, admin)

    template_response = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/import-template", headers=headers
    )
    assert template_response.status_code == 200
    template_workbook = load_workbook(BytesIO(template_response.content))
    assert [cell.value for cell in template_workbook["嘉宾导入"][1]][:7] == [
        "姓名", "手机号", "单位", "职务", "身份", "座位号", "学科"
    ]
    template_workbook.close()

    import_workbook = Workbook()
    worksheet = import_workbook.active
    worksheet.append(["姓名", "手机号", "单位", "职务", "身份", "座位号", "学科"])
    worksheet.append(["有效嘉宾", "13800000001", "第一学校", None, "教师", "B01", "数学"])
    worksheet.append(["缺手机号嘉宾", None, None, None, None, None, "语文"])
    upload = BytesIO()
    import_workbook.save(upload)
    import_workbook.close()
    import_response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests/import",
        headers=headers,
        files={"file": ("guests.xlsx", upload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_response.status_code == 200
    assert import_response.json()["imported_count"] == 1
    assert import_response.json()["errors"][0]["row_number"] == 3

    imported_guest = db.scalar(select(Guest).where(Guest.meeting_id == meeting.id, Guest.phone == "13800000001"))
    assert imported_guest is not None
    db.add(CheckIn(meeting_id=meeting.id, guest_id=imported_guest.id, staff_id=staff.id, method="manual"))
    db.commit()
    export_response = client.get(f"/api/admin/meetings/{meeting.id}/check-ins/export", headers=headers)
    assert export_response.status_code == 200
    export_workbook = load_workbook(BytesIO(export_response.content), data_only=True)
    exported_rows = list(export_workbook["签到明细"].iter_rows(values_only=True))
    export_workbook.close()
    assert exported_rows[1][1] == "有效嘉宾"
    assert exported_rows[1][7] == "已签到"
    assert exported_rows[1][10] == "staff-excel"


def test_admin_can_export_guest_information_and_status(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证嘉宾状态表包含正式嘉宾、未转化报名申请和对应状态。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示嘉宾状态表字段与管理列表口径一致。
    异常：当前函数不主动抛出业务异常；断言失败表示导出内容或去重规则不符合预期。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-guest-status-export")
    meeting = Meeting(title="嘉宾状态验证会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    subject_field = GuestField(meeting_id=meeting.id, label="学科", key="subject", field_type="text")
    db.add(subject_field)
    db.flush()
    approved_guest = Guest(
        meeting_id=meeting.id,
        name="已通过嘉宾",
        phone="13800000021",
        organization="第一学校",
        title="教师",
        source="self_registration",
        qr_token="guest-status-approved",
    )
    entered_guest = Guest(
        meeting_id=meeting.id,
        name="后台嘉宾",
        phone="13800000022",
        source="admin_import",
        qr_token="guest-status-entered",
    )
    db.add_all([approved_guest, entered_guest])
    db.flush()
    db.add(GuestValue(guest_id=approved_guest.id, field_id=subject_field.id, field_key="subject", value_text="数学"))
    db.add(CheckIn(meeting_id=meeting.id, guest_id=approved_guest.id, method="manual"))
    db.add_all([
        GuestApplication(
            meeting_id=meeting.id,
            name="待审核申请人",
            phone="13800000023",
            status="pending",
            values_json={"subject": "语文"},
        ),
        GuestApplication(
            meeting_id=meeting.id,
            name="已通过重复申请",
            phone="13800000024",
            status="approved",
            guest_id=approved_guest.id,
        ),
    ])
    db.commit()

    response = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/export",
        headers=auth_headers(db, admin),
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    rows = list(workbook["嘉宾状态"].iter_rows(values_only=True))
    workbook.close()
    assert rows[0] == (
        "记录ID", "姓名", "手机号", "单位", "职务", "身份", "座位号", "学科", "来源", "管理状态", "签到状态"
    )
    rows_by_name = {row[1]: row for row in rows[1:]}
    assert set(rows_by_name) == {"待审核申请人", "已通过嘉宾", "后台嘉宾"}
    assert rows_by_name["待审核申请人"][7:] == ("语文", "自主报名", "待审核", "—")
    assert rows_by_name["已通过嘉宾"][7:] == ("数学", "自主报名", "已通过", "已签到")
    assert rows_by_name["后台嘉宾"][8:] == ("后台导入", "已录入", "未签到")


def test_public_application_can_be_reviewed_into_guest(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证公开报名、重复保护、管理员查询和批准创建正式嘉宾。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示报名审核形成正式嘉宾的闭环可用。
    异常：当前函数不主动抛出业务异常；断言失败表示报名或审核规则异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-application")
    meeting = Meeting(
        title="开放报名会议",
        created_by_id=admin.id,
        status="published",
        end_time=datetime.now(timezone.utc) + timedelta(days=2),
    )
    db.add(meeting)
    db.flush()
    db.add_all([
        MeetingSetting(meeting_id=meeting.id, registration_enabled=True),
        MeetingAdmin(meeting_id=meeting.id, user_id=admin.id),
        GuestField(
            meeting_id=meeting.id,
            label="研究方向",
            key="research_area",
            field_type="text",
            required=True,
            visible_to_guest=True,
        ),
    ])
    db.commit()
    payload = {
        "name": "报名嘉宾",
        "phone": "13800000002",
        "organization": "第二学校",
        "values": {"research_area": "教育数字化"},
    }

    public_meeting_response = client.get(f"/api/meetings/{meeting.id}")
    assert public_meeting_response.status_code == 200
    assert public_meeting_response.json()["title"] == "开放报名会议"
    assert public_meeting_response.json()["guest_login_fields"] == ["name", "phone"]
    assert public_meeting_response.json()["registration_fields"] == [
        {"key": "name", "label": "姓名", "required": True},
        {"key": "phone", "label": "手机号", "required": True},
        {"key": "organization", "label": "单位", "required": False},
        {"key": "title", "label": "职务", "required": False},
        {"key": "research_area", "label": "研究方向", "required": True},
    ]

    submit_response = client.post(f"/api/meetings/{meeting.id}/guest-applications", json=payload)
    assert submit_response.status_code == 201
    application_id = submit_response.json()["id"]
    duplicate_response = client.post(f"/api/meetings/{meeting.id}/guest-applications", json=payload)
    assert duplicate_response.status_code == 422

    headers = auth_headers(db, admin)
    list_response = client.get(
        f"/api/admin/meetings/{meeting.id}/guest-applications?status=pending", headers=headers
    )
    assert list_response.status_code == 200
    assert [item["id"] for item in list_response.json()] == [application_id]
    review_response = client.patch(
        f"/api/admin/meetings/{meeting.id}/guest-applications/{application_id}",
        headers=headers,
        json={"status": "approved"},
    )
    assert review_response.status_code == 200
    assert review_response.json()["status"] == "approved"
    assert review_response.json()["guest_id"] is not None
    guest_response = client.get(f"/api/admin/meetings/{meeting.id}/guests", headers=headers)
    assert guest_response.status_code == 200
    assert guest_response.json()[0]["source"] == "self_registration"
    assert client.patch(
        f"/api/admin/meetings/{meeting.id}/guest-applications/{application_id}",
        headers=headers,
        json={"status": "rejected"},
    ).status_code == 422


def test_admin_resource_maintenance_and_cors_are_available(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证嘉宾、登录规则、多管理员维护以及前端跨域预检。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示补充资源 API 和默认开发跨域配置可用。
    异常：当前函数不主动抛出业务异常；断言失败表示补充接口未形成可用闭环。
    """
    client, db = client_and_session
    owner = create_user(db, "admin-resource-owner")
    other_admin = create_user(db, "admin-resource-other")
    meeting = Meeting(title="资源维护会议", created_by_id=owner.id, status="draft")
    db.add(meeting)
    db.flush()
    db.add_all([
        MeetingSetting(meeting_id=meeting.id),
        MeetingAdmin(meeting_id=meeting.id, user_id=owner.id),
        GuestField(meeting_id=meeting.id, label="学段", key="school_stage", field_type="text"),
    ])
    guest = Guest(meeting_id=meeting.id, name="待修改嘉宾", phone="13800000003", qr_token="resource-token")
    db.add(guest)
    db.commit()
    headers = auth_headers(db, owner)

    update_response = client.patch(
        f"/api/admin/meetings/{meeting.id}/guests/{guest.id}",
        headers=headers,
        json={"name": "已修改嘉宾", "values": {"school_stage": "高中"}},
    )
    assert update_response.status_code == 200
    assert update_response.json()["values"] == {"school_stage": "高中"}
    login_fields_response = client.put(
        f"/api/admin/meetings/{meeting.id}/guest-login-fields",
        headers=headers,
        json={"fields": ["phone", "name"]},
    )
    assert login_fields_response.json()["fields"] == ["name", "phone"]
    add_admin_response = client.post(
        f"/api/admin/meetings/{meeting.id}/admins",
        headers=headers,
        json={"username": other_admin.username},
    )
    assert add_admin_response.status_code == 200
    other_admin_headers = auth_headers(db, other_admin)
    assert client.get(f"/api/admin/meetings/{meeting.id}/admins", headers=other_admin_headers).status_code == 200
    assert client.delete(
        f"/api/admin/meetings/{meeting.id}/admins/{owner.id}", headers=headers
    ).status_code == 422
    assert client.delete(
        f"/api/admin/meetings/{meeting.id}/admins/{other_admin.id}", headers=headers
    ).status_code == 200
    assert client.get(f"/api/admin/meetings/{meeting.id}", headers=other_admin_headers).status_code == 404

    create_staff_response = client.post(
        f"/api/admin/meetings/{meeting.id}/staff",
        headers=headers,
        json={
            "username": "resource-staff",
            "display_name": "资源工作人员",
            "phone": "13700000009",
            "initial_password": "resource-pass-123",
        },
    )
    assert create_staff_response.status_code == 201
    staff_id = create_staff_response.json()["id"]
    patch_staff_response = client.patch(
        f"/api/admin/meetings/{meeting.id}/staff/{staff_id}",
        headers=headers,
        json={"display_name": "已修改工作人员", "is_active": False},
    )
    assert patch_staff_response.status_code == 200
    assert patch_staff_response.json()["is_active"] is False
    assert client.delete(
        f"/api/admin/meetings/{meeting.id}/staff/{staff_id}", headers=headers
    ).status_code == 200

    assert client.delete(
        f"/api/admin/meetings/{meeting.id}/guests/{guest.id}", headers=headers
    ).status_code == 200
    assert client.post(
        "/api/guest/sessions",
        json={"meeting_id": meeting.id, "name": "已修改嘉宾", "phone": guest.phone},
    ).status_code == 401

    cors_response = client.options(
        "/api/admin/meetings",
        headers={
            "Origin": "http://localhost:5173",
            "Access-Control-Request-Method": "GET",
            "Access-Control-Request-Headers": "authorization",
        },
    )
    assert cors_response.status_code == 200
    assert cors_response.headers["access-control-allow-origin"] == "http://localhost:5173"
    assert client.get("/api/admin/meetings").status_code == 401


def test_guest_deactivation_identity_and_dynamic_values_are_consistent(
    client_and_session: tuple[TestClient, Session],
) -> None:
    """验证停用嘉宾退出当前名单、身份可重新使用且动态必填字段遵守启用状态。

    入参：client_and_session 为测试客户端和数据库会话夹具。
    返回值：None：断言通过表示列表、统计、工作人员搜索、导出、身份唯一性和动态字段保存口径一致。
    异常：当前函数不主动抛出业务异常；断言失败表示嘉宾管理闭环存在数据口径回归。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-guest-consistency")
    staff = create_user(db, "staff-guest-consistency", role="staff")
    meeting = Meeting(title="嘉宾一致性会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add_all([
        MeetingAdmin(meeting_id=meeting.id, user_id=admin.id),
        StaffMeeting(meeting_id=meeting.id, user_id=staff.id),
        GuestField(
            meeting_id=meeting.id,
            label="饮食偏好",
            key="diet_preference",
            field_type="text",
            required=True,
            is_enabled=True,
        ),
        GuestField(
            meeting_id=meeting.id,
            label="历史字段",
            key="legacy_field",
            field_type="text",
            required=True,
            is_enabled=False,
        ),
    ])
    db.commit()
    admin_headers = auth_headers(db, admin)

    create_payload = {
        "name": "重复身份嘉宾",
        "phone": "13800000999",
        "values": {"diet_preference": "清淡"},
    }
    create_response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests",
        headers=admin_headers,
        json=create_payload,
    )
    assert create_response.status_code == 201
    guest_id = create_response.json()["id"]
    duplicate_response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests",
        headers=admin_headers,
        json=create_payload,
    )
    assert duplicate_response.status_code == 422
    assert duplicate_response.json()["detail"] == "当前会议已存在姓名和手机号相同的启用嘉宾。"

    db.add(CheckIn(meeting_id=meeting.id, guest_id=guest_id, staff_id=staff.id, method="manual"))
    db.commit()
    assert client.delete(
        f"/api/admin/meetings/{meeting.id}/guests/{guest_id}", headers=admin_headers
    ).status_code == 200
    assert client.get(f"/api/admin/meetings/{meeting.id}/guests", headers=admin_headers).json() == []

    summary_response = client.get(
        f"/api/admin/meetings/{meeting.id}/check-ins", headers=admin_headers
    )
    assert summary_response.status_code == 200
    assert summary_response.json() == {
        "total_guests": 0,
        "checked_in_count": 0,
        "unchecked_count": 0,
        "records": [],
    }
    staff_headers = auth_headers(db, staff)
    assert client.get(
        f"/api/staff/meetings/{meeting.id}/guests", headers=staff_headers
    ).json() == []

    check_in_export = client.get(
        f"/api/admin/meetings/{meeting.id}/check-ins/export", headers=admin_headers
    )
    check_in_workbook = load_workbook(BytesIO(check_in_export.content), read_only=True)
    assert check_in_workbook.active.max_row == 1
    check_in_workbook.close()
    guest_export = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/export", headers=admin_headers
    )
    guest_workbook = load_workbook(BytesIO(guest_export.content), read_only=True)
    assert guest_workbook.active.max_row == 1
    guest_workbook.close()

    recreate_response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests",
        headers=admin_headers,
        json={**create_payload, "values": {"diet_preference": "素食"}},
    )
    assert recreate_response.status_code == 201
    recreated_id = recreate_response.json()["id"]
    detail_response = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/{recreated_id}", headers=admin_headers
    )
    assert detail_response.status_code == 200
    assert detail_response.json()["values"] == {"diet_preference": "素食"}
