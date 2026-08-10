"""工作人员端同行人员登记 API 与导出测试。"""

from datetime import datetime, timedelta, timezone
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.access import MeetingAdmin, StaffMeeting
from app.models.guest import CheckIn, Guest
from app.models.meeting import Meeting
from app.models.user import User
from app.services.check_in_sessions import get_default_check_in_session


def prepare_staff_meeting(
    db: Session,
    admin: User,
    staff: User,
    title: str,
    end_time: datetime | None = None,
) -> Meeting:
    """创建已授权给工作人员的测试会议。

    入参：db 为数据库会话；admin 为会议创建管理员；staff 为工作人员；title 为会议名称，均必填；end_time 为会议结束时间，可为空。
    返回值：Meeting：已持久化的会议对象。
    异常：数据库写入失败时由 SQLAlchemy 抛出异常。
    """
    meeting = Meeting(title=title, created_by_id=admin.id, status="published", end_time=end_time)
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.add(StaffMeeting(meeting_id=meeting.id, user_id=staff.id))
    db.commit()
    return meeting


def register_companion(
    client: TestClient,
    meeting_id: int,
    staff_headers: dict[str, str],
    payload: dict[str, object],
):
    """调用工作人员同行登记接口并返回响应。

    入参：client 为测试客户端；meeting_id 为会议 ID；staff_headers 为工作人员请求头；payload 为登记数据，均必填。
    返回值：Response：同行登记接口响应。
    异常：当前函数不主动抛出业务异常。
    """
    return client.post(
        f"/api/staff/meetings/{meeting_id}/companions",
        headers=staff_headers,
        json=payload,
    )


def test_staff_registers_companion_successfully(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证工作人员可为主嘉宾登记同行人员并正确绑定。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示登记成功且来源、备注与绑定关系正确。
    异常：断言失败表示同行登记核心路径异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-companion")
    staff = create_user(db, "staff-companion", role="staff")
    meeting = prepare_staff_meeting(db, admin, staff, "同行登记会议")
    primary = Guest(meeting_id=meeting.id, name="李明", phone="13800001234", qr_token="primary-token")
    db.add(primary)
    db.commit()

    response = register_companion(
        client,
        meeting.id,
        auth_headers(db, staff),
        {
            "companion_of_id": primary.id,
            "name": "王秀兰",
            "phone": "13900005678",
            "organization": "华东师范大学",
            "companion_note": "家属",
        },
    )
    assert response.status_code == 201
    payload = response.json()
    assert payload["companion_of_id"] == primary.id
    assert payload["companion_of_name"] == "李明"
    assert payload["companion_note"] == "家属"
    assert payload["phone"] == "13900005678"

    companion = db.scalar(select(Guest).where(Guest.phone == "13900005678"))
    assert companion is not None
    assert companion.source == "companion_registration"
    assert companion.companion_of_id == primary.id
    assert companion.qr_token
    check_in = db.scalar(select(CheckIn).where(CheckIn.guest_id == companion.id))
    assert check_in is not None
    assert check_in.method == "manual"
    assert check_in.staff_id == staff.id

    search_response = client.get(
        f"/api/staff/meetings/{meeting.id}/guests",
        headers=auth_headers(db, staff),
        params={"query": "王秀兰"},
    )
    assert search_response.status_code == 200
    assert search_response.json()[0]["checked_in"] is True


def test_check_in_records_include_companion_binding(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证同行登记后签到记录列表返回完整绑定字段。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示签到记录携带同行绑定信息。
    异常：断言失败表示签到记录绑定字段异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-record-binding")
    staff = create_user(db, "staff-record-binding", role="staff")
    meeting = prepare_staff_meeting(db, admin, staff, "签到记录绑定会议")
    primary = Guest(meeting_id=meeting.id, name="李明", phone="13800001234", qr_token="binding-primary-token")
    db.add(primary)
    db.flush()
    default_session = get_default_check_in_session(db, meeting)
    # 主嘉宾先行手动签到，确保签到记录列表中同时存在主嘉宾与同行两条记录。
    db.add(
        CheckIn(
            meeting_id=meeting.id,
            session_id=default_session.id,
            guest_id=primary.id,
            staff_id=staff.id,
            method="scan",
        )
    )
    db.commit()

    register_response = register_companion(
        client,
        meeting.id,
        auth_headers(db, staff),
        {
            "companion_of_id": primary.id,
            "name": "王秀兰",
            "phone": "13900005678",
            "companion_note": "家属",
        },
    )
    assert register_response.status_code == 201

    records_response = client.get(
        f"/api/staff/meetings/{meeting.id}/check-ins",
        headers=auth_headers(db, staff),
    )
    assert records_response.status_code == 200
    records = records_response.json()
    assert len(records) == 2
    primary_record = next(record for record in records if not record["is_companion"])
    companion_record = next(record for record in records if record["is_companion"])
    assert primary_record["guest_name"] == "李明"
    assert primary_record["guest_phone"] == "13800001234"
    assert primary_record["companion_of_id"] is None
    assert companion_record["guest_name"] == "王秀兰"
    assert companion_record["guest_phone"] == "13900005678"
    assert companion_record["companion_of_id"] == primary.id
    assert companion_record["companion_of_name"] == "李明"
    assert companion_record["companion_note"] == "家属"
    assert companion_record["method"] == "manual"


def test_companion_rejects_invalid_or_other_meeting_primary(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证主嘉宾不存在或属于其他会议时拒绝登记。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示无效主嘉宾被拒绝。
    异常：断言失败表示主嘉宾校验缺失。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-invalid-primary")
    staff = create_user(db, "staff-invalid-primary", role="staff")
    meeting = prepare_staff_meeting(db, admin, staff, "无效主嘉宾会议")
    other_meeting = prepare_staff_meeting(db, admin, staff, "其他会议")
    other_guest = Guest(
        meeting_id=other_meeting.id,
        name="外会嘉宾",
        phone="13800001111",
        qr_token="other-meeting-token",
    )
    db.add(other_guest)
    db.commit()
    staff_headers = auth_headers(db, staff)

    missing_response = register_companion(
        client,
        meeting.id,
        staff_headers,
        {"companion_of_id": 99999, "name": "不存在", "phone": "13900000001"},
    )
    assert missing_response.status_code == 422

    other_response = register_companion(
        client,
        meeting.id,
        staff_headers,
        {"companion_of_id": other_guest.id, "name": "跨会嘉宾", "phone": "13900000002"},
    )
    assert other_response.status_code == 422
    assert "不属于当前会议" in other_response.json()["detail"]


def test_companion_rejects_chained_companion(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证同行嘉宾不能再作为主嘉宾登记其他同行人员。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示链式同行被拒绝。
    异常：断言失败表示链式同行校验缺失。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-chain")
    staff = create_user(db, "staff-chain", role="staff")
    meeting = prepare_staff_meeting(db, admin, staff, "链式同行会议")
    primary = Guest(meeting_id=meeting.id, name="主嘉宾", phone="13800002222", qr_token="chain-primary-token")
    db.add(primary)
    db.commit()
    staff_headers = auth_headers(db, staff)

    first_response = register_companion(
        client,
        meeting.id,
        staff_headers,
        {"companion_of_id": primary.id, "name": "同行甲", "phone": "13900002223"},
    )
    assert first_response.status_code == 201
    companion_id = first_response.json()["id"]

    chained_response = register_companion(
        client,
        meeting.id,
        staff_headers,
        {"companion_of_id": companion_id, "name": "同行乙", "phone": "13900002224"},
    )
    assert chained_response.status_code == 422
    assert "不能再作为主嘉宾" in chained_response.json()["detail"]


def test_companion_rejects_duplicate_identity(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证同会议内姓名和手机号相同的同行人员不能重复登记。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示重复身份被拒绝。
    异常：断言失败表示身份唯一校验缺失。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-duplicate-companion")
    staff = create_user(db, "staff-duplicate-companion", role="staff")
    meeting = prepare_staff_meeting(db, admin, staff, "重复同行会议")
    primary = Guest(meeting_id=meeting.id, name="主嘉宾", phone="13800003333", qr_token="dup-primary-token")
    db.add(primary)
    db.commit()
    staff_headers = auth_headers(db, staff)
    payload = {
        "companion_of_id": primary.id,
        "name": "重复同行",
        "phone": "13900003334",
    }

    assert register_companion(client, meeting.id, staff_headers, payload).status_code == 201
    duplicate_response = register_companion(client, meeting.id, staff_headers, payload)
    assert duplicate_response.status_code == 422
    assert "姓名和手机号相同" in duplicate_response.json()["detail"]


def test_staff_lists_companions_and_search_fields(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证同行列表可按主嘉宾筛选，且嘉宾搜索携带同行信息。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示列表过滤与搜索字段正确。
    异常：断言失败表示同行查询口径异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-list-companions")
    staff = create_user(db, "staff-list-companions", role="staff")
    meeting = prepare_staff_meeting(db, admin, staff, "同行列表会议")
    primary_a = Guest(meeting_id=meeting.id, name="嘉宾甲", phone="13800004444", qr_token="list-a-token")
    primary_b = Guest(meeting_id=meeting.id, name="嘉宾乙", phone="13800004445", qr_token="list-b-token")
    db.add_all([primary_a, primary_b])
    db.commit()
    staff_headers = auth_headers(db, staff)

    assert register_companion(
        client,
        meeting.id,
        staff_headers,
        {"companion_of_id": primary_a.id, "name": "同行丙", "phone": "13900004446"},
    ).status_code == 201
    assert register_companion(
        client,
        meeting.id,
        staff_headers,
        {"companion_of_id": primary_a.id, "name": "同行丁", "phone": "13900004447"},
    ).status_code == 201
    assert register_companion(
        client,
        meeting.id,
        staff_headers,
        {"companion_of_id": primary_b.id, "name": "同行戊", "phone": "13900004448"},
    ).status_code == 201

    filtered_response = client.get(
        f"/api/staff/meetings/{meeting.id}/companions",
        headers=staff_headers,
        params={"guest_id": primary_a.id},
    )
    assert filtered_response.status_code == 200
    filtered_payload = filtered_response.json()
    assert len(filtered_payload) == 2
    assert all(item["companion_of_id"] == primary_a.id for item in filtered_payload)
    assert all(item["companion_of_name"] == "嘉宾甲" for item in filtered_payload)

    guest_search_response = client.get(
        f"/api/staff/meetings/{meeting.id}/guests",
        headers=staff_headers,
    )
    assert guest_search_response.status_code == 200
    guests_by_name = {item["name"]: item for item in guest_search_response.json()}
    assert guests_by_name["嘉宾甲"]["companion_count"] == 2
    assert guests_by_name["嘉宾乙"]["companion_count"] == 1
    assert guests_by_name["同行丙"]["is_companion"] is True
    assert guests_by_name["同行丙"]["companion_of_name"] == "嘉宾甲"
    assert guests_by_name["嘉宾甲"]["is_companion"] is False


def test_exports_include_companion_columns(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证签到明细与嘉宾状态导出包含同行类型、陪同嘉宾和同行登记来源。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示导出表正确呈现同行信息。
    异常：断言失败表示导出字段缺失或来源标签错误。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-companion-export")
    staff = create_user(db, "staff-companion-export", role="staff")
    meeting = prepare_staff_meeting(db, admin, staff, "同行导出会议")
    primary = Guest(meeting_id=meeting.id, name="李明", phone="13800005555", qr_token="export-primary-token")
    db.add(primary)
    db.commit()
    staff_headers = auth_headers(db, staff)
    assert register_companion(
        client,
        meeting.id,
        staff_headers,
        {
            "companion_of_id": primary.id,
            "name": "王秀兰",
            "phone": "13900005556",
            "companion_note": "家属",
        },
    ).status_code == 201
    admin_headers = auth_headers(db, admin)

    check_in_export = client.get(
        f"/api/admin/meetings/{meeting.id}/check-ins/export",
        headers=admin_headers,
    )
    assert check_in_export.status_code == 200
    check_in_workbook = load_workbook(BytesIO(check_in_export.content), data_only=True)
    check_in_rows = list(check_in_workbook["签到明细"].iter_rows(values_only=True))
    check_in_workbook.close()
    assert check_in_rows[0][7] == "嘉宾类型"
    assert check_in_rows[0][8] == "陪同嘉宾"
    primary_row = next(row for row in check_in_rows[1:] if row[1] == "李明")
    companion_row = next(row for row in check_in_rows[1:] if row[1] == "王秀兰")
    assert primary_row[7] == "本人"
    assert primary_row[8] in (None, "")
    assert companion_row[7] == "同行人员"
    assert companion_row[8] == "李明"

    guest_export = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/export",
        headers=admin_headers,
    )
    assert guest_export.status_code == 200
    guest_workbook = load_workbook(BytesIO(guest_export.content), data_only=True)
    guest_rows = list(guest_workbook["嘉宾状态"].iter_rows(values_only=True))
    guest_workbook.close()
    header = guest_rows[0]
    type_index = header.index("嘉宾类型")
    companion_index = header.index("陪同嘉宾")
    source_index = header.index("来源")
    guest_companion_row = next(row for row in guest_rows[1:] if row[1] == "王秀兰")
    assert guest_companion_row[type_index] == "同行人员"
    assert guest_companion_row[companion_index] == "李明"
    assert guest_companion_row[source_index] == "同行登记"


def test_companion_requires_staff_permission(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证未授权工作人员不能登记同行人员。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示越权请求被拒绝。
    异常：断言失败表示权限校验缺失。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-companion-permission")
    staff = create_user(db, "staff-companion-permission", role="staff")
    other_staff = create_user(db, "staff-companion-other", role="staff")
    meeting = prepare_staff_meeting(db, admin, staff, "同行权限会议")
    primary = Guest(meeting_id=meeting.id, name="主嘉宾", phone="13800006666", qr_token="permission-primary-token")
    db.add(primary)
    db.commit()

    response = register_companion(
        client,
        meeting.id,
        auth_headers(db, other_staff),
        {"companion_of_id": primary.id, "name": "越权同行", "phone": "13900006667"},
    )
    assert response.status_code == 404


def test_companion_registration_fails_after_meeting_ends(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证会议结束后登记同行人员被拒绝，避免产生无法签到的孤立记录。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示会议结束后登记连带自动签到失败并整体拒绝。
    异常：断言失败表示会议结束校验缺失。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-ended-companion")
    staff = create_user(db, "staff-ended-companion", role="staff")
    meeting = prepare_staff_meeting(
        db,
        admin,
        staff,
        "已结束会议",
        end_time=datetime.now(timezone.utc) - timedelta(days=1),
    )
    primary = Guest(meeting_id=meeting.id, name="主嘉宾", phone="13800007777", qr_token="ended-primary-token")
    db.add(primary)
    db.commit()

    response = register_companion(
        client,
        meeting.id,
        auth_headers(db, staff),
        {"companion_of_id": primary.id, "name": "迟到同行", "phone": "13900007778"},
    )
    assert response.status_code == 422
    assert "会议已结束" in response.json()["detail"]
