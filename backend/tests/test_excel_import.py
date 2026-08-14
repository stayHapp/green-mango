"""Excel 导入导出 API 测试。

覆盖场景：模板下载、正常导入、错误行汇总、签到表导出、嘉宾状态导出等。
"""

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.access import MeetingAdmin
from app.models.guest import CheckIn, Guest, GuestField, GuestValue
from app.models.meeting import Meeting, MeetingSetting
from app.models.user import User
from app.services.check_in_sessions import get_default_check_in_session


def test_download_import_template_with_dynamic_fields(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证下载的导入模板包含动态字段列头。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示模板生成包含动态字段。
    异常：当前函数不主动抛出业务异常；断言失败表示模板生成逻辑异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-template")
    meeting = Meeting(title="模板会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.add(GuestField(meeting_id=meeting.id, label="学科", key="subject", field_type="text", required=True))
    db.commit()

    response = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/import-template",
        headers=auth_headers(db, admin),
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    header_row = [cell.value for cell in workbook["嘉宾导入"][1]][:7]
    assert header_row == ["姓名", "手机号", "单位", "职务", "身份", "座位号", "学科"]
    workbook.close()


def test_normal_import_succeeds(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证正常导入嘉宾成功。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示正常导入路径可用。
    异常：当前函数不主动抛出业务异常；断言失败表示导入逻辑异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-import")
    meeting = Meeting(title="导入会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.add(GuestField(meeting_id=meeting.id, label="学科", key="subject", field_type="text", required=True))
    db.commit()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["姓名", "手机号", "单位", "职务", "身份", "座位号", "学科"])
    worksheet.append(["有效嘉宾", "13800000001", "第一学校", "教师", "参会嘉宾", "B01", "数学"])
    upload = BytesIO()
    workbook.save(upload)
    workbook.close()

    response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests/import",
        headers=auth_headers(db, admin),
        files={"file": ("guests.xlsx", upload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    assert response.json()["imported_count"] == 1
    assert len(response.json()["errors"]) == 0

    imported_guest = db.scalar(select(Guest).where(Guest.meeting_id == meeting.id, Guest.phone == "13800000001"))
    assert imported_guest is not None
    assert imported_guest.name == "有效嘉宾"


def test_import_with_invalid_rows_reports_errors(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证导入含错误行时返回行级错误汇总；手机号留空不再报错。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示错误行被识别并报告。
    异常：当前函数不主动抛出业务异常；断言失败表示错误行处理逻辑异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-import-err")
    meeting = Meeting(title="错误导入会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.commit()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["姓名", "手机号", "单位", "职务", "身份", "座位号"])
    worksheet.append(["有效嘉宾", "13800000002", "第一学校", None, None, None])
    worksheet.append([None, "13800000003", None, None, None, None])
    upload = BytesIO()
    workbook.save(upload)
    workbook.close()

    response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests/import",
        headers=auth_headers(db, admin),
        files={"file": ("guests.xlsx", upload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    assert response.json()["imported_count"] == 1
    assert len(response.json()["errors"]) == 1
    assert response.json()["errors"][0]["row_number"] == 3


def test_check_in_export_produces_valid_xlsx(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证签到表导出为有效 XLSX 文件。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示签到导出生成正确文件。
    异常：当前函数不主动抛出业务异常；断言失败表示导出逻辑异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-export")
    staff = create_user(db, "staff-export", role="staff")
    meeting = Meeting(title="导出会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    guest = Guest(meeting_id=meeting.id, name="导出嘉宾", phone="13800000003", qr_token="export-token")
    db.add(guest)
    db.flush()
    default_session = get_default_check_in_session(db, meeting)
    db.add(CheckIn(meeting_id=meeting.id, session_id=default_session.id, guest_id=guest.id, staff_id=staff.id, method="manual"))
    db.commit()

    response = client.get(
        f"/api/admin/meetings/{meeting.id}/check-ins/export",
        headers=auth_headers(db, admin),
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    rows = list(workbook["签到明细"].iter_rows(values_only=True))
    workbook.close()
    assert len(rows) >= 2  # 至少表头 + 1 行数据
    assert rows[1][1] == "导出嘉宾"
    assert rows[1][7] == "本人"
    assert rows[1][8] in (None, "")
    assert rows[1][9] == "已签到"


def test_guest_status_export_produces_valid_xlsx(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证嘉宾状态导出为有效 XLSX 文件。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示嘉宾状态导出生成正确文件。
    异常：当前函数不主动抛出业务异常；断言失败表示导出逻辑异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-status-export")
    meeting = Meeting(title="状态导出会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    guest = Guest(
        meeting_id=meeting.id,
        name="状态嘉宾",
        phone="13800000004",
        source="admin_import",
        qr_token="status-export-token",
    )
    db.add(guest)
    db.commit()

    response = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/export",
        headers=auth_headers(db, admin),
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    rows = list(workbook["嘉宾状态"].iter_rows(values_only=True))
    workbook.close()
    assert len(rows) >= 2  # 至少表头 + 1 行数据
    # 查找嘉宾所在行
    guest_rows = [r for r in rows[1:] if r[1] == "状态嘉宾"]
    assert len(guest_rows) == 1


def test_exports_hide_disabled_guest_fields(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证导出文件按后台启用字段隐藏固定字段和动态字段。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示关闭座位号等字段后，签到明细和嘉宾状态表都不再导出对应列。
    异常：当前函数不主动抛出业务异常；断言失败表示停用字段可能在导出文件中泄露。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-hidden-export")
    staff = create_user(db, "staff-hidden-export", role="staff")
    meeting = Meeting(title="隐藏字段导出会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add_all([
        MeetingAdmin(meeting_id=meeting.id, user_id=admin.id),
        MeetingSetting(
            meeting_id=meeting.id,
            settings_json={
                "guest_enabled_fixed_fields": ["name", "phone", "organization"],
                "guest_registration_fields": ["name", "phone", "organization"],
                "guest_registration_required_fields": ["name", "phone"],
            },
        ),
    ])
    diet_field = GuestField(
        meeting_id=meeting.id,
        label="饮食偏好",
        key="diet_preference",
        field_type="text",
        is_enabled=True,
    )
    legacy_field = GuestField(
        meeting_id=meeting.id,
        label="历史字段",
        key="legacy_field",
        field_type="text",
        is_enabled=False,
    )
    db.add_all([diet_field, legacy_field])
    db.flush()
    guest = Guest(
        meeting_id=meeting.id,
        name="隐藏字段嘉宾",
        phone="13800000066",
        organization="测试单位",
        title="不应导出职务",
        tag="不应导出身份",
        seat="不应导出座位",
        qr_token="hidden-export-token",
    )
    db.add(guest)
    db.flush()
    default_session = get_default_check_in_session(db, meeting)
    db.add_all([
        GuestValue(
            guest_id=guest.id,
            field_id=diet_field.id,
            field_key=diet_field.key,
            value_text="清淡",
        ),
        GuestValue(
            guest_id=guest.id,
            field_id=legacy_field.id,
            field_key=legacy_field.key,
            value_text="历史值",
        ),
        CheckIn(
            meeting_id=meeting.id,
            session_id=default_session.id,
            guest_id=guest.id,
            staff_id=staff.id,
            method="manual",
        ),
    ])
    db.commit()

    check_in_response = client.get(
        f"/api/admin/meetings/{meeting.id}/check-ins/export",
        headers=auth_headers(db, admin),
    )
    assert check_in_response.status_code == 200
    check_in_workbook = load_workbook(BytesIO(check_in_response.content), data_only=True)
    check_in_rows = list(check_in_workbook["签到明细"].iter_rows(values_only=True))
    check_in_workbook.close()
    assert check_in_rows[0] == (
        "嘉宾ID",
        "姓名",
        "手机号",
        "单位",
        "嘉宾类型",
        "陪同嘉宾",
        "签到状态",
        "签到时间",
        "签到方式",
        "执行工作人员",
    )
    assert "不应导出座位" not in check_in_rows[1]
    assert "不应导出职务" not in check_in_rows[1]
    assert "不应导出身份" not in check_in_rows[1]

    guest_response = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/export",
        headers=auth_headers(db, admin),
    )
    assert guest_response.status_code == 200
    guest_workbook = load_workbook(BytesIO(guest_response.content), data_only=True)
    guest_rows = list(guest_workbook["嘉宾状态"].iter_rows(values_only=True))
    guest_workbook.close()
    assert guest_rows[0] == (
        "记录ID",
        "姓名",
        "手机号",
        "单位",
        "饮食偏好",
        "嘉宾类型",
        "陪同嘉宾",
        "来源",
        "管理状态",
        "签到状态",
    )
    assert "清淡" in guest_rows[1]
    assert "历史字段" not in guest_rows[0]
    assert "历史值" not in guest_rows[1]
    assert "不应导出座位" not in guest_rows[1]


def test_import_with_missing_phone_uses_default_phone(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证导入时手机号留空自动填充默认手机号。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示无手机号嘉宾可正常导入并使用默认号。
    异常：断言失败表示导入逻辑未按默认号处理。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-default-phone-import")
    meeting = Meeting(title="默认手机号导入会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.commit()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["姓名", "手机号", "单位", "职务", "身份", "座位号"])
    worksheet.append(["无手机号嘉宾", None, "示例学校", "教师", "参会嘉宾", "C02"])
    upload = BytesIO()
    workbook.save(upload)
    workbook.close()

    response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests/import",
        headers=auth_headers(db, admin),
        files={"file": ("guests.xlsx", upload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    assert response.json()["imported_count"] == 1
    assert response.json()["errors"] == []

    guest = db.scalar(select(Guest).where(Guest.name == "无手机号嘉宾"))
    assert guest is not None
    assert guest.phone == settings.default_guest_phone


def test_exports_show_not_provided_for_default_phone(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证导出文件把默认手机号显示为「未提供」。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示两种导出都不暴露默认手机号。
    异常：断言失败表示导出未按默认号转换。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-default-phone-export")
    meeting = Meeting(title="默认手机号导出会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    guest = Guest(
        meeting_id=meeting.id,
        name="无手机号嘉宾",
        phone=settings.default_guest_phone,
        source="admin_import",
        qr_token="default-phone-export-token",
    )
    db.add(guest)
    db.commit()

    check_in_response = client.get(
        f"/api/admin/meetings/{meeting.id}/check-ins/export",
        headers=auth_headers(db, admin),
    )
    assert check_in_response.status_code == 200
    check_in_workbook = load_workbook(BytesIO(check_in_response.content), data_only=True)
    check_in_rows = list(check_in_workbook["签到明细"].iter_rows(values_only=True))
    check_in_workbook.close()
    guest_check_in_row = [r for r in check_in_rows[1:] if r[1] == "无手机号嘉宾"]
    assert len(guest_check_in_row) == 1
    assert guest_check_in_row[0][2] == "未提供"

    guest_response = client.get(
        f"/api/admin/meetings/{meeting.id}/guests/export",
        headers=auth_headers(db, admin),
    )
    assert guest_response.status_code == 200
    guest_workbook = load_workbook(BytesIO(guest_response.content), data_only=True)
    guest_rows = list(guest_workbook["嘉宾状态"].iter_rows(values_only=True))
    guest_workbook.close()
    guest_export_row = [r for r in guest_rows[1:] if r[1] == "无手机号嘉宾"]
    assert len(guest_export_row) == 1
    assert guest_export_row[0][2] == "未提供"


def test_import_overwrites_existing_guest_and_reports_summary(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证重复导入同名同手机号嘉宾覆盖更新并返回统计摘要。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示覆盖更新、签到保留与统计字段正确。
    异常：断言失败表示导入覆盖或统计逻辑异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-import-overwrite")
    meeting = Meeting(title="覆盖导入会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.commit()
    admin_headers = auth_headers(db, admin)

    def build_upload(phone: str, organization: str) -> BytesIO:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["姓名", "手机号", "单位", "职务", "身份", "座位号"])
        worksheet.append(["覆盖嘉宾", phone, organization, "教师", "参会嘉宾", "D01"])
        upload = BytesIO()
        workbook.save(upload)
        workbook.close()
        return upload

    first_response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests/import",
        headers=admin_headers,
        files={"file": ("guests.xlsx", build_upload("13900010001", "第一学校").getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert first_response.status_code == 200
    assert first_response.json()["imported_count"] == 1
    assert first_response.json()["updated_count"] == 0

    guest = db.scalar(select(Guest).where(Guest.name == "覆盖嘉宾"))
    assert guest is not None
    guest_id = guest.id

    overwrite_response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests/import",
        headers=admin_headers,
        files={"file": ("guests.xlsx", build_upload("13900010001", "第二学校").getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert overwrite_response.status_code == 200
    assert overwrite_response.json()["imported_count"] == 0
    assert overwrite_response.json()["updated_count"] == 1
    assert overwrite_response.json()["default_phone_count"] == 0

    db.expire_all()
    updated_guest = db.get(Guest, guest_id)
    assert updated_guest is not None
    assert updated_guest.organization == "第二学校"

    default_phone_response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests/import",
        headers=admin_headers,
        files={"file": ("guests.xlsx", build_upload("", "第三学校").getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert default_phone_response.status_code == 200
    assert default_phone_response.json()["updated_count"] == 1
    assert default_phone_response.json()["default_phone_count"] == 1


def test_import_accepts_extra_columns_and_missing_phone_column(
    client_and_session: tuple[TestClient, Session],
    create_user,
    auth_headers,
) -> None:
    """验证导入允许多余列且可省略手机号列。

    入参：client_and_session 为测试客户端和数据库会话夹具；create_user 为创建用户辅助函数；auth_headers 为请求头辅助函数。
    返回值：None：断言通过表示表头顺序、多余列与缺手机号列均不阻断导入。
    异常：断言失败表示导入表头兼容逻辑异常。
    """
    client, db = client_and_session
    admin = create_user(db, "admin-import-flexible")
    meeting = Meeting(title="灵活导入会议", created_by_id=admin.id, status="published")
    db.add(meeting)
    db.flush()
    db.add(MeetingAdmin(meeting_id=meeting.id, user_id=admin.id))
    db.commit()
    admin_headers = auth_headers(db, admin)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["邮箱", "姓名", "单位", "备注"])
    worksheet.append(["zhang@example.com", "灵活嘉宾", "示例学校", "无手机号列"])
    upload = BytesIO()
    workbook.save(upload)
    workbook.close()

    response = client.post(
        f"/api/admin/meetings/{meeting.id}/guests/import",
        headers=admin_headers,
        files={"file": ("guests.xlsx", upload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    assert response.json()["imported_count"] == 1
    assert response.json()["default_phone_count"] == 1
    assert response.json()["errors"] == []

    guest = db.scalar(select(Guest).where(Guest.name == "灵活嘉宾"))
    assert guest is not None
    assert guest.phone == settings.default_guest_phone
    assert guest.organization == "示例学校"
